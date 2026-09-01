from __future__ import annotations

from pathlib import Path
from datetime import datetime, timezone
import json
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .classifier import RULESET_VERSION
from .models import EvaluationRecord, EvaluationSummary, LabeledExample


def load_labeled_examples(path: str | Path) -> list[LabeledExample]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        examples: list[LabeledExample] = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            system_number = "" if not row or row[0] is None else str(row[0]).strip()
            expected = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
            if row_number == 1 and _is_header(system_number, expected):
                continue
            if not system_number and not expected:
                continue
            if not system_number or not expected:
                raise ValueError(
                    f"Row {row_number} must contain both a system number and expected system type"
                )
            examples.append(LabeledExample(row_number, system_number, expected))
    finally:
        workbook.close()
    if not examples:
        raise ValueError("The workbook contains no labeled examples")
    return examples


def export_evaluation_report(path: str | Path, summary: EvaluationSummary) -> None:
    output_path = Path(path)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Summary"
    overview.append(["Ruleset Version", RULESET_VERSION])
    overview.append(["Total Records", len(summary.records)])
    overview.append([])
    overview.append(["Evaluation Status", "Count"])
    for status, count in sorted(summary.counts.items()):
        overview.append([status, count])

    _write_records_sheet(workbook, "Results", summary.records)
    _write_records_sheet(
        workbook,
        "Mismatches",
        (record for record in summary.records if record.evaluation_status == "MISMATCH"),
    )
    _write_records_sheet(
        workbook,
        "Needs Review",
        (
            record
            for record in summary.records
            if record.evaluation_status not in {"MATCH", "MISMATCH"}
        ),
    )

    _style_sheet(overview)
    workbook.save(output_path)
    manifest_path = output_path.with_suffix(".json")
    manifest_path.write_text(
        json.dumps(
            {
                "ruleset_version": RULESET_VERSION,
                "generated_utc": datetime.now(timezone.utc).isoformat(),
                "total_records": len(summary.records),
                "status_counts": summary.counts,
                "records": [
                    {
                        "source_row": record.source_row,
                        "system_number": record.system_number,
                        "expected_system_type": record.expected_system_type,
                        "predicted_system_type": record.predicted_system_type,
                        "evaluation_status": record.evaluation_status,
                        "build_type": record.build_type,
                        "matched_rule_ids": record.rule_ids,
                        "decision_evidence": record.evidence,
                        "warnings_or_failure_reason": record.warnings,
                    }
                    for record in summary.records
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _write_records_sheet(
    workbook: Workbook,
    title: str,
    records: Iterable[EvaluationRecord],
) -> None:
    sheet = workbook.create_sheet(title)
    headers = [
        "Source Row",
        "System Number",
        "Expected System Type",
        "Predicted System Type",
        "Evaluation Status",
        "Build Type",
        "Matched Rule IDs",
        "Decision Evidence",
        "Warnings / Failure Reason",
    ]
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                record.source_row,
                record.system_number,
                record.expected_system_type,
                record.predicted_system_type,
                record.evaluation_status,
                record.build_type,
                record.rule_ids,
                record.evidence,
                record.warnings,
            ]
        )
    _style_sheet(sheet)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="176B63")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        maximum = max((len(str(cell.value or "")) for cell in column), default=0)
        sheet.column_dimensions[letter].width = min(max(maximum + 2, 12), 70)
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _is_header(first: str, second: str) -> bool:
    first_key = first.lower().replace(" ", "_")
    second_key = second.lower().replace(" ", "_")
    return first_key in {"system_number", "part_number", "system"} and second_key in {
        "expected_system_type",
        "system_type",
        "correct_system_type",
        "type",
    }
