from __future__ import annotations

from dataclasses import dataclass
from enum import StrEnum
import logging
import os
from pathlib import Path
import re
import shutil
import uuid
import warnings
from typing import Mapping

from openpyxl import load_workbook

from .classifier import RULESET_VERSION
from .models import (
    DecisionStatus,
    SystemClassification,
    SystemNumberInput,
)
from .parser import parse_system_number
from .templates import SYSTEM_TYPE_TO_WD_TEMPLATE, TEMPLATE_MAP_VERSION


LOGGER = logging.getLogger(__name__)
_HEADER_SCAN_ROWS = 25
_HEADER_SCAN_COLUMNS = 50
EVIDENCE_SHEET_NAME = "AMAT Match Evidence"
WORKFLOW_VERSION = "2026.08.28.1"


class WorkflowMode(StrEnum):
    SYSTEM_TYPE = "SYSTEM_TYPE"
    WD_TEMPLATE = "WD_TEMPLATE"


class VerificationOutcome(StrEnum):
    CONFIRMED = "CONFIRMED"
    CORRECTED = "CORRECTED"
    REJECTED = "REJECTED"


@dataclass(frozen=True)
class VerificationFeedback:
    outcome: VerificationOutcome
    corrected_system_type: str = ""
    notes: str = ""


@dataclass(frozen=True)
class WorkbookLayout:
    source_path: Path
    sheet_name: str
    system_column: int
    output_column: int
    header_row: int | None
    output_header: str | None
    inputs: tuple[SystemNumberInput, ...]


@dataclass(frozen=True)
class CellValueUpdate:
    row: int
    column: int
    value: str


@dataclass(frozen=True)
class MatchEvidence:
    source_row: int
    system_number: str
    mode: str
    build_type: str
    classification_status: str
    proposed_system_type: str
    output_system_type: str
    output_wd_template: str
    matched_rule_ids: str
    decision_evidence: str
    warnings: str
    user_verification: str
    user_corrected_type: str
    user_notes: str
    requirements_action: str
    ruleset_version: str
    template_map_version: str
    workflow_version: str


@dataclass(frozen=True)
class WorkbookUpdatePlan:
    layout: WorkbookLayout
    updates: tuple[CellValueUpdate, ...]
    evidence: tuple[MatchEvidence, ...]
    written_count: int
    blank_count: int


def load_quote_request_layout(path: str | Path) -> WorkbookLayout:
    source_path = _validate_source_path(path)
    workbook = _load_for_reading(source_path)
    try:
        match = _find_header_pair(workbook, "system number", "system type")
        if match is None:
            raise ValueError(
                "No worksheet contains System Number and System Type headers "
                f"on the same row within the first {_HEADER_SCAN_ROWS} rows"
            )
        sheet, header_row, system_column, output_column = match
        inputs = _read_inputs(sheet, header_row + 1, system_column)
        if not inputs:
            raise ValueError(
                f"Worksheet {sheet.title!r} contains no system numbers below "
                f"cell {sheet.cell(header_row, system_column).coordinate}"
            )
        return WorkbookLayout(
            source_path=source_path,
            sheet_name=sheet.title,
            system_column=system_column,
            output_column=output_column,
            header_row=header_row,
            output_header=None,
            inputs=tuple(inputs),
        )
    finally:
        workbook.close()


def load_template_input_layout(path: str | Path) -> WorkbookLayout:
    source_path = _validate_source_path(path)
    workbook = _load_for_reading(source_path)
    try:
        header = _find_system_number_header(workbook)
        if header is not None:
            sheet, header_row, system_column = header
            output_column = system_column + 1
            adjacent_header = sheet.cell(header_row, output_column).value
            normalized_adjacent = _normalize_header(adjacent_header)
            allowed_headers = {"", "wd template", "template", "system template"}
            if normalized_adjacent not in allowed_headers:
                raise ValueError(
                    f"Cannot use adjacent column {output_column} on worksheet "
                    f"{sheet.title!r}; its header is {adjacent_header!r}"
                )
            inputs = _read_inputs(sheet, header_row + 1, system_column)
            if not inputs:
                raise ValueError(
                    f"Worksheet {sheet.title!r} contains no system numbers below "
                    f"the header in row {header_row}"
                )
            if not normalized_adjacent:
                _reject_populated_adjacent_cells(sheet, inputs, output_column)
            return WorkbookLayout(
                source_path=source_path,
                sheet_name=sheet.title,
                system_column=system_column,
                output_column=output_column,
                header_row=header_row,
                output_header="WD Template",
                inputs=tuple(inputs),
            )

        sheet, system_column, first_row = _find_headerless_system_column(workbook)
        inputs = _read_inputs(sheet, first_row, system_column)
        output_column = system_column + 1
        _reject_populated_adjacent_cells(sheet, inputs, output_column)
        return WorkbookLayout(
            source_path=source_path,
            sheet_name=sheet.title,
            system_column=system_column,
            output_column=output_column,
            header_row=None,
            output_header=None,
            inputs=tuple(inputs),
        )
    finally:
        workbook.close()


def build_update_plan(
    layout: WorkbookLayout,
    classifications: list[SystemClassification],
    mode: WorkflowMode,
    feedback_by_row: Mapping[int, VerificationFeedback] | None = None,
) -> WorkbookUpdatePlan:
    feedback_by_row = feedback_by_row or {}
    expected_rows = [item.source_row for item in layout.inputs]
    actual_rows = [item.source_row for item in classifications]
    if actual_rows != expected_rows:
        raise ValueError(
            "Classification results are incomplete or no longer match workbook order"
        )

    updates: list[CellValueUpdate] = []
    evidence_rows: list[MatchEvidence] = []
    written_count = 0
    for classification in classifications:
        decision = classification.decision
        feedback = feedback_by_row.get(classification.source_row)
        output_system_type = _approved_system_type(decision, feedback)
        output_wd_template = (
            SYSTEM_TYPE_TO_WD_TEMPLATE[output_system_type]
            if output_system_type and mode == WorkflowMode.WD_TEMPLATE
            else ""
        )
        value = (
            output_system_type
            if mode == WorkflowMode.SYSTEM_TYPE
            else output_wd_template
        )
        if value:
            written_count += 1
        updates.append(
            CellValueUpdate(classification.source_row, layout.output_column, value)
        )
        user_verification = "NOT_REQUIRED"
        corrected_type = ""
        user_notes = ""
        requirements_action = ""
        if decision.status == DecisionStatus.VERIFICATION_REQUIRED:
            user_verification = "PENDING" if feedback is None else feedback.outcome.value
            if feedback is not None:
                corrected_type = feedback.corrected_system_type
                user_notes = feedback.notes
                requirements_action = {
                    VerificationOutcome.CONFIRMED: "CONFIDENCE_EXAMPLE",
                    VerificationOutcome.CORRECTED: "RULE_CORRECTION",
                    VerificationOutcome.REJECTED: "RULE_REVIEW",
                }[feedback.outcome]
            else:
                requirements_action = "USER_VERIFICATION_REQUIRED"
        evidence_rows.append(
            MatchEvidence(
                source_row=classification.source_row,
                system_number=classification.system_number,
                mode=mode.value,
                build_type=classification.build_type,
                classification_status=decision.status.value,
                proposed_system_type=decision.predicted_system_type,
                output_system_type=output_system_type,
                output_wd_template=output_wd_template,
                matched_rule_ids="; ".join(decision.rule_ids),
                decision_evidence="\n".join(decision.evidence),
                warnings="\n".join(decision.warnings),
                user_verification=user_verification,
                user_corrected_type=corrected_type,
                user_notes=user_notes,
                requirements_action=requirements_action,
                ruleset_version=RULESET_VERSION,
                template_map_version=TEMPLATE_MAP_VERSION,
                workflow_version=WORKFLOW_VERSION,
            )
        )

    return WorkbookUpdatePlan(
        layout=layout,
        updates=tuple(updates),
        evidence=tuple(evidence_rows),
        written_count=written_count,
        blank_count=len(updates) - written_count,
    )


def write_value_only_workbook_copy(
    plan: WorkbookUpdatePlan,
    output_path: str | Path,
) -> Path:
    source = plan.layout.source_path.resolve()
    output = Path(output_path).expanduser().resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("Output file must use the .xlsx extension")
    if output == source:
        raise ValueError("Output path must be different from the source workbook")
    if not output.parent.exists():
        raise ValueError(f"Output folder does not exist: {output.parent}")

    temp = output.with_name(f".{output.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    shutil.copy2(source, temp)
    try:
        _write_cells_with_excel(plan, temp)
        os.replace(temp, output)
    finally:
        try:
            temp.unlink(missing_ok=True)
        except OSError:
            LOGGER.warning("Unable to remove temporary workbook %s", temp.name)
    return output


def _write_cells_with_excel(plan: WorkbookUpdatePlan, workbook_path: Path) -> None:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError(
            "Microsoft Excel automation support is unavailable. Install pywin32."
        ) from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(
            str(workbook_path),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
        )
        sheet = workbook.Worksheets(plan.layout.sheet_name)
        if plan.layout.header_row is not None and plan.layout.output_header is not None:
            sheet.Cells(plan.layout.header_row, plan.layout.output_column).Value2 = (
                plan.layout.output_header
            )
        for update in plan.updates:
            sheet.Cells(update.row, update.column).Value2 = update.value or None
        _write_evidence_sheet(workbook, plan.evidence)
        sheet.Activate()
        workbook.Save()
    except Exception as exc:
        raise RuntimeError(f"Excel could not write the output workbook: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                LOGGER.exception("Unable to close the temporary Excel workbook")
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                LOGGER.exception("Unable to close the temporary Excel process")
        pythoncom.CoUninitialize()


def _approved_system_type(decision, feedback: VerificationFeedback | None) -> str:
    if decision.status == DecisionStatus.CLASSIFIED:
        if feedback is not None:
            raise ValueError("Verification feedback was supplied for a classified result")
        return decision.predicted_system_type
    if decision.status != DecisionStatus.VERIFICATION_REQUIRED:
        if feedback is not None:
            raise ValueError(
                f"Verification feedback was supplied for {decision.status.value}"
            )
        return ""
    if feedback is None or feedback.outcome == VerificationOutcome.REJECTED:
        return ""
    if feedback.outcome == VerificationOutcome.CONFIRMED:
        return decision.predicted_system_type
    corrected = feedback.corrected_system_type.strip()
    if corrected not in SYSTEM_TYPE_TO_WD_TEMPLATE:
        raise ValueError(f"Corrected system type is not canonical: {corrected!r}")
    return corrected


def _write_evidence_sheet(workbook, evidence: tuple[MatchEvidence, ...]) -> None:
    sheet = _get_evidence_sheet(workbook)
    sheet.Cells.Clear()
    headers = (
        "Source Row",
        "System Number",
        "Mode",
        "Build Type",
        "Classification Status",
        "Proposed System Type",
        "Output System Type",
        "Output WD Template",
        "Matched Rule IDs",
        "Decision Evidence",
        "Warnings",
        "User Verification",
        "User Corrected Type",
        "User Notes",
        "Requirements Action",
        "Ruleset Version",
        "Template Map Version",
        "Workflow Version",
    )
    rows = tuple(
        (
            row.source_row,
            row.system_number,
            row.mode,
            row.build_type,
            row.classification_status,
            row.proposed_system_type,
            row.output_system_type,
            row.output_wd_template,
            row.matched_rule_ids,
            row.decision_evidence,
            row.warnings,
            row.user_verification,
            row.user_corrected_type,
            row.user_notes,
            row.requirements_action,
            row.ruleset_version,
            row.template_map_version,
            row.workflow_version,
        )
        for row in evidence
    )
    matrix = (headers,) + rows
    target = sheet.Range(
        sheet.Cells(1, 1),
        sheet.Cells(len(matrix), len(headers)),
    )
    target.Value2 = matrix
    header = sheet.Range(sheet.Cells(1, 1), sheet.Cells(1, len(headers)))
    header.Font.Bold = True
    header.Interior.Color = 0x784E1F
    target.AutoFilter(Field=1)
    target.Columns.AutoFit()
    for column in range(1, len(headers) + 1):
        column_range = sheet.Columns(column)
        if column_range.ColumnWidth > 60:
            column_range.ColumnWidth = 60
    for column in (9, 10, 11, 14):
        sheet.Columns(column).WrapText = True
    if sheet.Index < workbook.Worksheets.Count:
        sheet.Move(None, workbook.Worksheets(workbook.Worksheets.Count))


def _get_evidence_sheet(workbook):
    try:
        existing = workbook.Worksheets(EVIDENCE_SHEET_NAME)
    except Exception:
        existing = None
    if existing is not None:
        if existing.Cells(1, 1).Value2 in {None, "Source Row"}:
            return existing
        name = _unique_sheet_name(workbook, EVIDENCE_SHEET_NAME)
    else:
        name = EVIDENCE_SHEET_NAME
    sheet = workbook.Worksheets.Add()
    sheet.Name = name
    return sheet


def _unique_sheet_name(workbook, base: str) -> str:
    existing = {
        str(workbook.Worksheets(index).Name).casefold()
        for index in range(1, workbook.Worksheets.Count + 1)
    }
    for sequence in range(2, 100):
        suffix = f" {sequence}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate.casefold() not in existing:
            return candidate
    raise RuntimeError("Unable to allocate an evidence worksheet name")


def _validate_source_path(path: str | Path) -> Path:
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() != ".xlsx":
        raise ValueError("Input file must use the .xlsx extension")
    if not source.is_file():
        raise ValueError(f"Input workbook not found: {source}")
    return source


def _load_for_reading(path: Path):
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*extension is not supported and will be removed.*",
                category=UserWarning,
                module="openpyxl.worksheet._reader",
            )
            return load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ValueError(f"Unable to read workbook {path.name}: {exc}") from exc


def _find_header_pair(workbook, first_header: str, second_header: str):
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, _HEADER_SCAN_ROWS) + 1):
            found: dict[str, int] = {}
            for column in range(1, min(sheet.max_column, _HEADER_SCAN_COLUMNS) + 1):
                normalized = _normalize_header(sheet.cell(row, column).value)
                if normalized in {first_header, second_header}:
                    found[normalized] = column
            if first_header in found and second_header in found:
                return sheet, row, found[first_header], found[second_header]
    return None


def _find_system_number_header(workbook):
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, _HEADER_SCAN_ROWS) + 1):
            for column in range(1, min(sheet.max_column, _HEADER_SCAN_COLUMNS) + 1):
                if _normalize_header(sheet.cell(row, column).value) in {
                    "system number",
                    "system numbers",
                }:
                    return sheet, row, column
    return None


def _find_headerless_system_column(workbook):
    candidates: list[tuple[int, object, int, int]] = []
    for sheet in workbook.worksheets:
        for column in range(1, min(sheet.max_column, _HEADER_SCAN_COLUMNS) + 1):
            valid_rows = [
                row
                for row in range(1, sheet.max_row + 1)
                if _is_valid_system_number(sheet.cell(row, column).value)
            ]
            if valid_rows:
                candidates.append((len(valid_rows), sheet, column, min(valid_rows)))
    if not candidates:
        raise ValueError(
            "No System Number header or column containing valid system numbers was found"
        )
    candidates.sort(key=lambda item: item[0], reverse=True)
    if len(candidates) > 1 and candidates[0][0] == candidates[1][0]:
        raise ValueError(
            "More than one headerless column contains the same number of valid "
            "system numbers; add a System Number header to identify the input column"
        )
    _, sheet, column, first_row = candidates[0]
    return sheet, column, first_row


def _read_inputs(sheet, first_row: int, system_column: int) -> list[SystemNumberInput]:
    inputs: list[SystemNumberInput] = []
    for row in range(first_row, sheet.max_row + 1):
        value = sheet.cell(row, system_column).value
        if value is None or not str(value).strip():
            continue
        inputs.append(SystemNumberInput(row, str(value).strip()))
    return inputs


def _reject_populated_adjacent_cells(sheet, inputs, output_column: int) -> None:
    populated = [
        sheet.cell(item.source_row, output_column).coordinate
        for item in inputs
        if sheet.cell(item.source_row, output_column).value not in {None, ""}
    ]
    if populated:
        preview = ", ".join(populated[:5])
        suffix = "..." if len(populated) > 5 else ""
        raise ValueError(
            "The adjacent output column contains existing data at "
            f"{preview}{suffix}; add or rename its header to WD Template only if "
            "those cells may be overwritten"
        )


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def _is_valid_system_number(value) -> bool:
    if value is None:
        return False
    return parse_system_number(str(value).strip()).valid
