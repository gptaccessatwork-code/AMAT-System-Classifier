import json
from pathlib import Path
import unittest
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from system_type_identifier.models import EvaluationRecord, EvaluationSummary
from system_type_identifier.workbook import export_evaluation_report, load_labeled_examples


class WorkbookTests(unittest.TestCase):
    def scratch_path(self, suffix: str) -> Path:
        path = Path.cwd() / f"test_{uuid4().hex}{suffix}"
        self.addCleanup(lambda: path.unlink(missing_ok=True))
        return path

    def test_headerless_labeled_workbook(self):
        path = self.scratch_path("_batch.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["C02130-EY2-GP1", "EPI SINGLE CLUSTER"])
        sheet.append(["708973-XA3T-GPA", "ETCH SYM3 AP (XA)"])
        workbook.save(path)

        examples = load_labeled_examples(path)
        self.assertEqual(len(examples), 2)
        self.assertEqual(examples[0].source_row, 1)

    def test_report_and_manifest_export(self):
        path = self.scratch_path("_evaluation.xlsx")
        manifest_path = path.with_suffix(".json")
        self.addCleanup(lambda: manifest_path.unlink(missing_ok=True))
        summary = EvaluationSummary(
            [
                EvaluationRecord(
                    1,
                    "C02130-EY2-GP1",
                    "EPI SINGLE CLUSTER",
                    "EPI SINGLE CLUSTER",
                    "MATCH",
                    "NORMAL",
                    "SYS-EY1-EY2",
                    "Product family is EY2",
                    "",
                )
            ]
        )
        export_evaluation_report(path, summary)

        workbook = load_workbook(path, read_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary", "Results", "Mismatches", "Needs Review"])
        workbook.close()
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["status_counts"], {"MATCH": 1})
        self.assertEqual(
            manifest["records"][0]["decision_evidence"],
            "Product family is EY2",
        )


if __name__ == "__main__":
    unittest.main()
