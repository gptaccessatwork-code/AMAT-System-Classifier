from pathlib import Path
import unittest
from unittest.mock import patch
from uuid import uuid4

from openpyxl import Workbook

from system_type_identifier.models import (
    ClassificationDecision,
    DecisionStatus,
    SystemClassification,
)
from system_type_identifier.workflows import (
    VerificationFeedback,
    VerificationOutcome,
    WorkflowMode,
    build_update_plan,
    load_quote_request_layout,
    load_template_input_layout,
    write_value_only_workbook_copy,
)


class WorkbookWorkflowTests(unittest.TestCase):
    def scratch_path(self, suffix: str = ".xlsx") -> Path:
        path = Path.cwd() / f"test_{uuid4().hex}{suffix}"
        self.addCleanup(lambda: path.unlink(missing_ok=True))
        return path

    def test_quote_layout_finds_headers_and_rows(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Quote Requests"
        sheet["D6"] = "System Number"
        sheet["H6"] = "System Type"
        sheet["D7"] = "707844-XA3-GPA"
        sheet["H7"] = "OLD VALUE"
        sheet["D9"] = "707281-XG3-GPLL"
        workbook.save(path)

        layout = load_quote_request_layout(path)

        self.assertEqual(layout.sheet_name, "Quote Requests")
        self.assertEqual(layout.header_row, 6)
        self.assertEqual(layout.system_column, 4)
        self.assertEqual(layout.output_column, 8)
        self.assertEqual(
            [(item.source_row, item.system_number) for item in layout.inputs],
            [(7, "707844-XA3-GPA"), (9, "707281-XG3-GPLL")],
        )

    def test_template_layout_accepts_header_and_empty_adjacent_column(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["C3"] = "System Number"
        sheet["C4"] = "707844-XA3-GPA"
        workbook.save(path)

        layout = load_template_input_layout(path)

        self.assertEqual(layout.header_row, 3)
        self.assertEqual(layout.system_column, 3)
        self.assertEqual(layout.output_column, 4)
        self.assertEqual(layout.output_header, "WD Template")

    def test_template_layout_accepts_headerless_column(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["B2"] = "707844-XA3-GPA"
        sheet["B3"] = "707281-XG3-GPLL"
        workbook.save(path)

        layout = load_template_input_layout(path)

        self.assertIsNone(layout.header_row)
        self.assertEqual(layout.system_column, 2)
        self.assertEqual(layout.output_column, 3)
        self.assertEqual([item.source_row for item in layout.inputs], [2, 3])

    def test_template_layout_rejects_unrelated_adjacent_column(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "System Number"
        sheet["B1"] = "Customer"
        sheet["A2"] = "707844-XA3-GPA"
        sheet["B2"] = "Example Customer"
        workbook.save(path)

        with self.assertRaisesRegex(ValueError, "adjacent column"):
            load_template_input_layout(path)

    def test_update_plan_requires_verification_and_blanks_review_rows(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["D6"] = "System Number"
        sheet["H6"] = "System Type"
        sheet["D7"] = "710001-XP-GP"
        sheet["D8"] = "BAD"
        workbook.save(path)
        layout = load_quote_request_layout(path)
        classifications = [
            SystemClassification(
                7,
                "710001-XP-GP",
                "NORMAL",
                ClassificationDecision(
                    DecisionStatus.VERIFICATION_REQUIRED,
                    "ETCH NEXTGEN 2 CHAMBER",
                ),
            ),
            SystemClassification(
                8,
                "BAD",
                "",
                ClassificationDecision(DecisionStatus.EXCLUDED_INVALID_FORMAT),
            ),
        ]

        unverified = build_update_plan(
            layout,
            classifications,
            WorkflowMode.SYSTEM_TYPE,
        )
        self.assertEqual([update.value for update in unverified.updates], ["", ""])

        verified = build_update_plan(
            layout,
            classifications,
            WorkflowMode.WD_TEMPLATE,
            feedback_by_row={
                7: VerificationFeedback(VerificationOutcome.CONFIRMED)
            },
        )
        self.assertEqual(
            [update.value for update in verified.updates],
            ["SGP_TEMPLATE_AMAT_NEXTGEN_2", ""],
        )
        self.assertEqual(verified.evidence[0].user_verification, "CONFIRMED")
        self.assertEqual(
            verified.evidence[0].requirements_action,
            "CONFIDENCE_EXAMPLE",
        )
        self.assertEqual(verified.evidence[1].classification_status, "EXCLUDED_INVALID_FORMAT")

    def test_corrected_verification_uses_corrected_type_and_template(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "System Number"
        sheet["A2"] = "710001-XP-GP"
        workbook.save(path)
        layout = load_template_input_layout(path)
        classifications = [
            SystemClassification(
                2,
                "710001-XP-GP",
                "NORMAL",
                ClassificationDecision(
                    DecisionStatus.VERIFICATION_REQUIRED,
                    "ETCH NEXTGEN 2 CHAMBER",
                    ("BOM-NEXTGEN-CHAMBER-COUNT",),
                    ("Chamber designators indicate A and B",),
                ),
            )
        ]

        plan = build_update_plan(
            layout,
            classifications,
            WorkflowMode.WD_TEMPLATE,
            feedback_by_row={
                2: VerificationFeedback(
                    VerificationOutcome.CORRECTED,
                    corrected_system_type="ETCH NEXTGEN 3 CHAMBER",
                    notes="BOM also contains chamber C",
                )
            },
        )

        self.assertEqual(plan.updates[0].value, "SGP_TEMPLATE_AMAT_NEXTGEN_3")
        evidence = plan.evidence[0]
        self.assertEqual(evidence.output_system_type, "ETCH NEXTGEN 3 CHAMBER")
        self.assertEqual(evidence.user_verification, "CORRECTED")
        self.assertEqual(evidence.user_notes, "BOM also contains chamber C")
        self.assertEqual(evidence.requirements_action, "RULE_CORRECTION")

    def test_rejected_verification_stays_blank_and_becomes_rule_review(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "System Number"
        sheet["A2"] = "710001-XP-GP"
        workbook.save(path)
        layout = load_template_input_layout(path)
        classifications = [
            SystemClassification(
                2,
                "710001-XP-GP",
                "NORMAL",
                ClassificationDecision(
                    DecisionStatus.VERIFICATION_REQUIRED,
                    "ETCH NEXTGEN 2 CHAMBER",
                ),
            )
        ]

        plan = build_update_plan(
            layout,
            classifications,
            WorkflowMode.SYSTEM_TYPE,
            feedback_by_row={
                2: VerificationFeedback(
                    VerificationOutcome.REJECTED,
                    notes="Unable to confirm chamber count",
                )
            },
        )

        self.assertEqual(plan.updates[0].value, "")
        self.assertEqual(plan.evidence[0].requirements_action, "RULE_REVIEW")
        self.assertEqual(plan.evidence[0].user_verification, "REJECTED")

    def test_noncanonical_correction_is_rejected(self):
        path = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "System Number"
        sheet["A2"] = "710001-XP-GP"
        workbook.save(path)
        layout = load_template_input_layout(path)
        classifications = [
            SystemClassification(
                2,
                "710001-XP-GP",
                "NORMAL",
                ClassificationDecision(
                    DecisionStatus.VERIFICATION_REQUIRED,
                    "ETCH NEXTGEN 2 CHAMBER",
                ),
            )
        ]

        with self.assertRaisesRegex(ValueError, "not canonical"):
            build_update_plan(
                layout,
                classifications,
                WorkflowMode.SYSTEM_TYPE,
                feedback_by_row={
                    2: VerificationFeedback(
                        VerificationOutcome.CORRECTED,
                        corrected_system_type="MADE UP TYPE",
                    )
                },
            )

    def test_writer_uses_copy_and_rejects_source_overwrite(self):
        source = self.scratch_path()
        output = self.scratch_path()
        workbook = Workbook()
        sheet = workbook.active
        sheet["D1"] = "System Number"
        sheet["H1"] = "System Type"
        sheet["D2"] = "707844-XA3-GPA"
        workbook.save(source)
        layout = load_quote_request_layout(source)
        classifications = [
            SystemClassification(
                2,
                "707844-XA3-GPA",
                "NORMAL",
                ClassificationDecision(
                    DecisionStatus.CLASSIFIED,
                    "ETCH SYM3 AP (XA)",
                ),
            )
        ]
        plan = build_update_plan(
            layout,
            classifications,
            WorkflowMode.SYSTEM_TYPE,
        )

        with patch("system_type_identifier.workflows._write_cells_with_excel"):
            result = write_value_only_workbook_copy(plan, output)
        self.assertEqual(result, output.resolve())
        self.assertTrue(output.is_file())
        self.assertEqual(
            list(output.parent.glob(f".{output.stem}.*.tmp.xlsx")),
            [],
        )
        with self.assertRaisesRegex(ValueError, "different from the source"):
            write_value_only_workbook_copy(plan, source)


if __name__ == "__main__":
    unittest.main()
